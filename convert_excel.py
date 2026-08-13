#!/usr/bin/env python3
"""
Convert Excel to GoodsBuy backup JSON format.

Usage:
    python convert_excel.py input.xlsx output.json

Excel format (single sheet, no multi-sheet):
    - Header row with column names (Chinese or English keys accepted)
    - Supported columns:
        名称/name, 种类/category, 类型/type, IP/系列名/ipName, 制品系列/seriesName,
        角色/标签/characterTag, 备注/remark, 购买渠道/purchaseChannel, 购买店铺/purchaseShop,
        购买日期/purchaseDate, 购买价格/purchasePrice, 购买数量/purchaseQuantity,
        购买运费/purchaseShipping, 预期售价/expectedPrice, 售出价格/sellPrice,
        售出数量/sellQuantity, 售出运费/sellShipping, 包邮/isFreeShipping,
        售出日期/sellDate, 买家信息/buyerInfo, 售出备注/sellRemark,
        状态/status, 仓储状态/storageStatus
    - Status values must be English enum keys (OWNED, SOLD, etc.)
    - Storage status values must be English enum keys (IN_STOCK, IN_TRANSIT, etc.)
"""

import sys
import json
import re
from pathlib import Path
from datetime import datetime, date

try:
    from openpyxl import load_workbook
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ── Status / storage mapping ────────────────────────────────────────────
STATUS_MAP = {
    '已出': 'SOLD',
    '犹豫出': 'HESITATING_SELL',
    '待出': 'LISTED',
    '持有': 'OWNED',
    '待补': 'PENDING_TAIL',
    '待邮': 'PENDING_SHIPPING_FEE',
    '待发货': 'PENDING_SEND',
    '运输中': 'IN_TRANSIT_ORDER',
    '赠品': 'GIFT',
    '遗失': 'LOST',
    '损坏': 'LOST',
}

STORAGE_MAP = {
    '现货': 'IN_STOCK',
    '在途': 'IN_TRANSIT',
    '团长': 'GROUP_STORAGE',
    '代购处囤货': 'AGENT_STORAGE',
}

VALID_STATUSES = {'PENDING_TAIL','PENDING_SHIPPING_FEE','PENDING_SEND',
                  'IN_TRANSIT_ORDER','OWNED','HESITATING_SELL',
                  'LISTED','SOLD','GIFT','LOST'}
VALID_STORAGE  = {'IN_STOCK','IN_TRANSIT','GROUP_STORAGE','AGENT_STORAGE'}

# ── Header alias map ────────────────────────────────────────────────────
HEADER_ALIASES = {
    '名称': 'name', 'name': 'name',
    '种类': 'category', 'category': 'category',
    '类型': 'type', 'type': 'type',
    'ip系列名': 'ipName', 'ip': 'ipName', 'ip名称': 'ipName',
    '系列名': 'seriesName', '系列': 'seriesName', '制品系列': 'seriesName',
    '角色': 'characterTag', '标签': 'characterTag', '角色标签': 'characterTag',
    '备注': 'remark', 'remark': 'remark',
    '购买渠道': 'purchaseChannel', 'purchaseChannel': 'purchaseChannel', '渠道': 'purchaseChannel',
    '购买店铺': 'purchaseShop', 'purchaseShop': 'purchaseShop',
    '购买日期': 'purchaseDate', 'purchaseDate': 'purchaseDate', 'date': 'purchaseDate',
    '购买价格': 'purchasePrice', 'purchasePrice': 'purchasePrice', '原价': 'purchasePrice',
    '收价': 'purchasePrice', '入手价': 'purchasePrice',
    '购买数量': 'purchaseQuantity', 'purchaseQuantity': 'purchaseQuantity', '数量': 'purchaseQuantity',
    '购买运费': 'purchaseShipping', 'purchaseShipping': 'purchaseShipping', '运费': 'purchaseShipping',
    '预期售价': 'expectedPrice', 'expectedPrice': 'expectedPrice', '意愿': 'expectedPrice',
    '出价': 'sellPrice', '售价': 'sellPrice', 'sellPrice': 'sellPrice',
    '售出数量': 'sellQuantity', 'sellQuantity': 'sellQuantity',
    '售出运费': 'sellShipping', 'sellShipping': 'sellShipping',
    '包邮': 'isFreeShipping', 'isFreeShipping': 'isFreeShipping', 'freeShipping': 'isFreeShipping',
    '售出日期': 'sellDate', 'sellDate': 'sellDate',
    '买家信息': 'buyerInfo', 'buyerInfo': 'buyerInfo',
    '售出备注': 'sellRemark', 'sellRemark': 'sellRemark',
    '状态': 'status', 'status': 'status',
    '仓储状态': 'storageStatus', 'storageStatus': 'storageStatus',
}


def resolve_header(raw: str) -> str | None:
    """Return the canonical field name for a header string, or None."""
    s = str(raw).strip()
    s_lower = s.lower()
    if s in HEADER_ALIASES:
        return HEADER_ALIASES[s]
    for key, val in HEADER_ALIASES.items():
        k_lower = key.lower()
        if k_lower in s_lower or s_lower in k_lower:
            return val
    return None


def parse_date(val):
    """Return epoch ms or None."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        # Excel serial date
        if val > 40000:
            dt = date(1899, 12, 30) + __import__('datetime').timedelta(days=int(val))
            return int(dt.replace(tzinfo=None).timestamp() * 1000)
        return int(val)
    s = str(val).strip()
    m = re.match(r'(\d{4})[-/年](\d{1,2})[-/月]?(\d{1,2})?[-日]?', s)
    if m:
        yr, mo, dy = int(m.group(1)), int(m.group(2)), int(m.group(3) or 1)
        try:
            return int(datetime(yr, mo, dy).timestamp() * 1000)
        except ValueError:
            return None
    return None


def parse_price(val):
    """Parse price – take first numeric token from a possibly-multi-value string."""
    if val is None or str(val).strip() == '':
        return 0.0
    text = str(val).replace(',', '').replace('¥', '').replace('￥', '')
    text = re.sub(r'[，、/+\-]', ' ', text)
    nums = re.findall(r'\d+(?:\.\d+)?', text)
    if nums:
        try:
            return float(nums[0])
        except ValueError:
            return 0.0
    return 0.0


def parse_quantity(val):
    if val is None or str(val).strip() == '':
        return 1
    nums = re.findall(r'\d+', str(val))
    if nums:
        try:
            return int(nums[0])
        except ValueError:
            return 1
    return 1


def parse_bool(val):
    if val is None:
        return False
    s = str(val).strip().lower()
    return s in ('true', 'yes', 'y', '是', '1', 't')


def parse_status(val):
    if val is None or str(val).strip() == '':
        return 'OWNED'
    s = str(val).strip()
    if s in VALID_STATUSES:
        return s
    for key, status in STATUS_MAP.items():
        if key in s:
            return status
    return 'OWNED'


def parse_storage(val):
    if val is None or str(val).strip() == '':
        return 'IN_STOCK'
    s = str(val).strip()
    if s in VALID_STORAGE:
        return s
    for key, status in STORAGE_MAP.items():
        if key in s:
            return status
    return 'IN_STOCK'


def main():
    if len(sys.argv) < 3:
        print("Usage: python convert_excel.py input.xlsx output.json")
        sys.exit(1)

    input_path  = Path(sys.argv[1])
    output_path = Path(sys.argv[2])

    if not input_path.exists():
        print(f"Error: File not found: {input_path}")
        sys.exit(1)

    print(f"Reading: {input_path}")
    wb = load_workbook(input_path, data_only=True)
    ws = wb.active  # use first sheet

    # Auto-detect header row: scan rows 1-5 for a row with recognized column names
    HEADER_SEARCH_ROWS = 5
    header_row_idx = None
    for test_r in range(1, HEADER_SEARCH_ROWS + 1):
        test_headers = [ws.cell(row=test_r, column=c).value for c in range(1, 30)]
        known_count = sum(1 for h in test_headers if resolve_header(str(h)) is not None)
        if known_count >= 3:
            header_row_idx = test_r
            break

    if header_row_idx is None:
        print("Error: Could not find header row. Expected at least 3 recognized column names.")
        sys.exit(1)

    print(f"  Header row: {header_row_idx}")
    header_row = [ws.cell(row=header_row_idx, column=c).value for c in range(1, 30)]
    col_map = {}
    for i, h in enumerate(header_row):
        if h is None:
            continue
        field = resolve_header(h)
        if field:
            col_map[field] = i

    print(f"  Mapped columns: {list(col_map.keys())}")

    def cell(row, field):
        idx = col_map.get(field)
        if idx is None:
            return None
        v = ws.cell(row=row, column=idx + 1).value
        return str(v).strip() if v is not None else None

    records = []
    now_ms  = int(datetime.now().timestamp() * 1000)
    next_id = 1

    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        name = cell(row_idx, 'name')
        if not name:
            continue  # skip empty rows

        purchase_date = parse_date(cell(row_idx, 'purchaseDate')) or now_ms
        sell_price_raw = cell(row_idx, 'sellPrice')
        sell_price = parse_price(sell_price_raw) if sell_price_raw else None
        expected = parse_price(cell(row_idx, 'expectedPrice')) if cell(row_idx, 'expectedPrice') else (sell_price or 0.0)

        r = {
            'id': next_id,
            'name': name,
            'category': cell(row_idx, 'category') or '',
            'type': cell(row_idx, 'type') or '官方',
            'ipName': cell(row_idx, 'ipName') or '',
            'seriesName': cell(row_idx, 'seriesName') or '',
            'characterTag': cell(row_idx, 'characterTag') or '',
            'remark': cell(row_idx, 'remark') or '',
            'purchaseChannel': cell(row_idx, 'purchaseChannel') or '',
            'purchaseShop': cell(row_idx, 'purchaseShop') or '',
            'purchaseDate': purchase_date,
            'purchasePrice': parse_price(cell(row_idx, 'purchasePrice')),
            'purchaseQuantity': parse_quantity(cell(row_idx, 'purchaseQuantity')),
            'purchaseShipping': parse_price(cell(row_idx, 'purchaseShipping')),
            'expectedPrice': expected,
            'sellPrice': sell_price,
            'sellQuantity': int(parse_price(cell(row_idx, 'sellQuantity'))) if cell(row_idx, 'sellQuantity') else None,
            'sellShipping': parse_price(cell(row_idx, 'sellShipping')) if cell(row_idx, 'sellShipping') else None,
            'isFreeShipping': parse_bool(cell(row_idx, 'isFreeShipping')),
            'sellDate': parse_date(cell(row_idx, 'sellDate')),
            'buyerInfo': cell(row_idx, 'buyerInfo') or None,
            'sellRemark': cell(row_idx, 'sellRemark') or None,
            'status': parse_status(cell(row_idx, 'status')),
            'storageStatus': parse_storage(cell(row_idx, 'storageStatus')),
            'imageFilenames': [],
            'createdAt': now_ms,
            'updatedAt': now_ms,
        }
        next_id += 1
        records.append(r)

    manifest = {
        'version': 1,
        'timestamp': now_ms,
        'collectibles': records,
    }

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)

    print(f"\nDone! {len(records)} records written to: {output_path}")


if __name__ == '__main__':
    main()
