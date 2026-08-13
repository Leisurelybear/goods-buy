#!/usr/bin/env python3
"""Generate an Excel template for GoodsBuy collectible import."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = Workbook()
ws = wb.active
ws.title = "藏品导入"

# Columns matching CollectibleRecord fields
COLUMNS = [
    ("name",            "名称",               40),
    ("category",        "种类",               15),
    ("type",            "类型",               12),
    ("ipName",          "IP/系列名",          20),
    ("seriesName",      "制品系列",           20),
    ("characterTag",    "角色/标签",          20),
    ("remark",          "备注",               30),
    ("purchaseChannel", "购买渠道",           15),
    ("purchaseShop",    "购买店铺",           20),
    ("purchaseDate",    "购买日期",           15),
    ("purchasePrice",   "购买价格",           12),
    ("purchaseQuantity","购买数量",           10),
    ("purchaseShipping","购买运费",           12),
    ("expectedPrice",   "预期售价",           12),
    ("sellPrice",       "售出价格",           12),
    ("sellQuantity",    "售出数量",           10),
    ("sellShipping",    "售出运费",           12),
    ("isFreeShipping",  "包邮",               8),
    ("sellDate",        "售出日期",           15),
    ("buyerInfo",       "买家信息",           20),
    ("sellRemark",      "售出备注",           30),
    ("status",          "状态",               12),
    ("storageStatus",   "仓储状态",           12),
]

HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT   = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
ALT_FILL     = PatternFill("solid", fgColor="EEF3F9")
WARN_FILL    = PatternFill("solid", fgColor="FFF3CD")
THIN_BORDER  = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

STATUS_OPTIONS = "PENDING_TAIL,PENDING_SHIPPING_FEE,PENDING_SEND,IN_TRANSIT_ORDER,OWNED,HESITATING_SELL,LISTED,SOLD,GIFT,LOST"
STATUS_CN      = "待补尾款,待补邮,待发货,运输中,已拥有,犹豫出售,已挂出,已售出,赠品,遗失"
STORAGE_OPTIONS = "IN_STOCK,IN_TRANSIT,GROUP_STORAGE,AGENT_STORAGE"
STORAGE_CN      = "现货,在途,团长囤货,代购处囤货"

# ── Row 1: titles ──────────────────────────────────────────────────────
ws.merge_cells("A1:W1")
title_cell = ws["A1"]
title_cell.value = "GoodsBuy 藏品导入模板"
title_cell.font = Font(name="微软雅黑", bold=True, size=16, color="1F4E79")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

# ── Row 2: subtitle ────────────────────────────────────────────────────
ws.merge_cells("A2:W2")
sub = ws["A2"]
sub.value = ("请在此表格中填写藏品信息，然后使用 convert_excel.py 转换为 JSON 导入到 App。"
             "状态/仓储状态列请使用英文枚举值。")
sub.font = Font(name="微软雅黑", size=10, color="555555")
sub.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[2].height = 22

# ── Row 3: blank spacer ────────────────────────────────────────────────
ws.row_dimensions[3].height = 6

# ── Row 4: column headers ──────────────────────────────────────────────
for col_idx, (key, label, width) in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=4, column=col_idx, value=label)
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = CENTER
    cell.border    = THIN_BORDER
    ws.column_dimensions[get_column_letter(col_idx)].width = width
ws.row_dimensions[4].height = 30

# ── Rows 5-24: empty data rows ─────────────────────────────────────────
for r in range(5, 25):
    ws.row_dimensions[r].height = 22
    for c in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=r, column=c, value="")
        cell.alignment = LEFT
        cell.border    = THIN_BORDER
        if r % 2 == 1:
            cell.fill = ALT_FILL

# ── Column X: status help ──────────────────────────────────────────────
ws.merge_cells("X4:X4")
ws["X4"] = "状态说明"
ws["X4"].font      = Font(bold=True, size=11, color="1F4E79")
ws["X4"].alignment = CENTER

status_notes = [
    ("PENDING_TAIL",   "待补尾款"),
    ("PENDING_SHIPPING_FEE", "待补邮"),
    ("PENDING_SEND",   "待发货"),
    ("IN_TRANSIT_ORDER","运输中"),
    ("OWNED",          "已拥有"),
    ("HESITATING_SELL","犹豫出售"),
    ("LISTED",         "已挂出"),
    ("SOLD",           "已售出"),
    ("GIFT",           "赠品/付邮送"),
    ("LOST",           "遗失/损坏"),
]
for i, (en, cn) in enumerate(status_notes):
    r = 5 + i
    ws.cell(row=r, column=24, value=en).alignment = LEFT
    ws.cell(row=r, column=25, value=cn).alignment = LEFT

ws.merge_cells("X14:X14")
ws["X14"] = "仓储状态说明"
ws["X14"].font = Font(bold=True, size=11, color="1F4E79")
ws["X14"].alignment = CENTER
for i, (en, cn) in enumerate([
    ("IN_STOCK",     "现货"),
    ("IN_TRANSIT",   "在途"),
    ("GROUP_STORAGE","团长囤货"),
    ("AGENT_STORAGE","代购处囤货"),
]):
    r = 15 + i
    ws.cell(row=r, column=24, value=en).alignment = LEFT
    ws.cell(row=r, column=25, value=cn).alignment = LEFT

# ── Row 26: blank ──────────────────────────────────────────────────────
ws.row_dimensions[26].height = 10

# ── Row 27: example ────────────────────────────────────────────────────
ws.merge_cells("A27:W27")
ws["A27"].value = "示例行（可直接删除）"
ws["A27"].font = Font(name="微软雅黑", bold=True, size=11, color="2E7D32")
ws.row_dimensions[27].height = 24

example = [
    "进击的巨人立牌-A", "手办/模型", "官方", "进击的巨人", "剧场版", "利威尔", "",
    "淘宝", "漫潮旗舰店", datetime(2025, 3, 15), 89.0, 1, 12.0, 120.0,
    None, None, None, False, None, None, "", "OWNED", "IN_STOCK",
]
for c, v in enumerate(example, start=1):
    cell = ws.cell(row=28, column=c, value=v)
    cell.fill = PatternFill("solid", fgColor="E8F5E9")
    cell.border = THIN_BORDER
    cell.alignment = LEFT

# ── Free shipping note ─────────────────────────────────────────────────
ws["Y29"] = "说明："
ws["Y29"].font = Font(bold=True, size=10)
ws.merge_cells("Z29:AD29")
ws["Z29"].value = ("包邮：填写 true/false；日期格式 YYYY-MM-DD；状态和仓储状态必须填写英文枚举值；"
                   "空白字段留空即可。")

out_path = "G:/Coding_Project/IdeaProjects/goods_collector/goods_buy_import_template.xlsx"
wb.save(out_path)
print(f"Template saved: {out_path}")
